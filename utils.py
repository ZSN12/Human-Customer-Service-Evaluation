# utils.py
import pandas as pd
import io
import openpyxl
import logging

def estimate_rows(uploaded_file) -> int:
    """估算上传文件的行数（不全部读入内存）"""
    try:
        content = uploaded_file.read(1024 * 1024)  # 读前1MB
        uploaded_file.seek(0)
        if uploaded_file.name.lower().endswith('.csv'):
            # 粗略估算：按换行符计数，减去表头
            lines = content.decode('utf-8', errors='ignore').count('\n')
            return max(0, lines)
        else:
            # Excel文件使用openpyxl的read_only模式读取真实行数
            try:
                wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                ws = wb.active
                row_count = ws.max_row
                wb.close()
                uploaded_file.seek(0)
                return row_count if row_count else 50000  # 默认50000
            except Exception as e:
                logging.warning(f"无法准确估算Excel行数: {e}")
                uploaded_file.seek(0)
                return 50000  # 返回一个较大的默认值
    except:
        return 0

def read_any_file(uploaded_file, chunk_size=None):
    """读取上传的CSV或Excel文件，返回DataFrame"""
    # 检查 uploaded_file 是字符串还是文件对象
    if isinstance(uploaded_file, str):
        # 如果是字符串，作为文件路径处理
        file_path = uploaded_file
        ext = file_path.lower().split('.')[-1]
        
        try:
            if ext == 'csv':
                if chunk_size:
                    # 分块读取并合并
                    chunks = []
                    for chunk in pd.read_csv(file_path, encoding='utf-8-sig', chunksize=chunk_size):
                        chunks.append(chunk)
                    if chunks:
                        return pd.concat(chunks, ignore_index=True)
                    else:
                        return pd.DataFrame()
                else:
                    return pd.read_csv(file_path, encoding='utf-8-sig')
            elif ext in ['xlsx', 'xls']:
                # 行数较少，直接使用pandas读取
                return pd.read_excel(file_path, engine='openpyxl')
            else:
                return pd.DataFrame()
        except UnicodeDecodeError:
            if ext == 'csv':
                if chunk_size:
                    # 分块读取并合并
                    chunks = []
                    for chunk in pd.read_csv(file_path, encoding='gbk', chunksize=chunk_size):
                        chunks.append(chunk)
                    if chunks:
                        return pd.concat(chunks, ignore_index=True)
                    else:
                        return pd.DataFrame()
                else:
                    return pd.read_csv(file_path, encoding='gbk')
            else:
                raise
    else:
        # 如果是文件对象，按原逻辑处理
        ext = uploaded_file.name.lower().split('.')[-1]
        try:
            if ext == 'csv':
                if chunk_size:
                    # 分块读取并合并
                    chunks = []
                    for chunk in pd.read_csv(uploaded_file, encoding='utf-8-sig', chunksize=chunk_size):
                        chunks.append(chunk)
                    if chunks:
                        return pd.concat(chunks, ignore_index=True)
                    else:
                        return pd.DataFrame()
                else:
                    return pd.read_csv(uploaded_file, encoding='utf-8-sig')
            elif ext in ['xlsx', 'xls']:
                # 先估算行数
                row_count = estimate_rows(uploaded_file)
                uploaded_file.seek(0)
                
                from config import settings
                # 如果行数超过警告阈值，使用read_only模式
                if row_count > settings.max_row_warning:
                    # 提示用户将大Excel文件转换为CSV以获得更好的性能
                    import streamlit as st
                    st.warning("检测到大型Excel文件，建议转换为CSV格式以获得更好的性能和内存使用效率")
                    
                    # 仍然尝试读取，但使用分块方式
                    import openpyxl
                    
                    # 读取工作簿
                    wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # 读取表头
                    header = []
                    for cell in ws[1]:
                        header.append(cell.value)
                    
                    # 分块读取数据
                    chunk_size = 10000
                    chunks = []
                    current_chunk = []
                    
                    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        current_chunk.append(row)
                        if len(current_chunk) >= chunk_size:
                            chunks.append(pd.DataFrame(current_chunk, columns=header))
                            current_chunk = []
                    
                    # 处理剩余数据
                    if current_chunk:
                        chunks.append(pd.DataFrame(current_chunk, columns=header))
                    
                    wb.close()
                    
                    # 合并分块
                    if chunks:
                        df = pd.concat(chunks, ignore_index=True)
                        return df
                    else:
                        return pd.DataFrame(columns=header)
                else:
                    # 行数较少，直接使用pandas读取
                    return pd.read_excel(uploaded_file, engine='openpyxl')
            else:
                return pd.DataFrame()
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            if ext == 'csv':
                if chunk_size:
                    # 分块读取并合并
                    chunks = []
                    for chunk in pd.read_csv(uploaded_file, encoding='gbk', chunksize=chunk_size):
                        chunks.append(chunk)
                    if chunks:
                        return pd.concat(chunks, ignore_index=True)
                    else:
                        return pd.DataFrame()
                else:
                    return pd.read_csv(uploaded_file, encoding='gbk')
            else:
                raise